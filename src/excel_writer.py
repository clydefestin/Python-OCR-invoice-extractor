"""
excel_writer.py

Purpose:
    Export extracted receipt information into Excel.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class ExcelWriter:

    def __init__(self):

        self.columns = [
            "Page",
            "Receipt No.",
            "Doctor Name",
            "PRC License",
            "Hospital",
            "Date",
            "Patient Name",
            "Total Amount (PHP)",
            "Signature"
        ]

    def export(self, records, output_path):

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Doctor Receipts"

        # Header
        for col, header in enumerate(self.columns, start=1):

            cell = sheet.cell(row=1, column=col)

            cell.value = header
            cell.font = Font(bold=True)

            cell.fill = PatternFill(
                fill_type="solid",
                start_color="D9EAD3"
            )

        # Data
        for row, record in enumerate(records, start=2):

            for col, header in enumerate(self.columns, start=1):

                sheet.cell(
                    row=row,
                    column=col
                ).value = record.get(header, "")

        # Auto width
        for column_cells in sheet.columns:

            length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column_cells
            )

            sheet.column_dimensions[
                column_cells[0].column_letter
            ].width = length + 3

        output_path = Path(output_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        workbook.save(output_path)

        print(f"\n[SUCCESS] Excel exported to:\n{output_path}")